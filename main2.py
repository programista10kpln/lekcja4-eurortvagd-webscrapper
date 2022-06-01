import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
cell_number = 1

page_counter = 1
pages = 0

while True:
    try:
        pages = int(input('ile stron chcesz przeskanować?\n'))
        if pages > 1:
            break
        else:
            print('zła wartość')
            continue
    except ValueError:
        print('zła wartość')
        continue

for page_counter in range(1, pages):

    url = 'https://www.euro.com.pl/telefony-komorkowe,_xiaomi,strona-' + str(page_counter) + '.bhtml'

    site = requests.get(url)
    soup = BeautifulSoup(site.content, 'html.parser')

    xiaomi_s = soup.find_all('div', class_="product-row")

    for xiaomi in xiaomi_s:
        name = xiaomi.find('h2', class_='product-name').text.strip()
        number = xiaomi.find('div', class_='selenium-product-code').next_element.text.strip()
        attrs_name_html = xiaomi.find_all('span', class_='attribute-name')
        attrs_name = []
        for attr in attrs_name_html:
            attrs_name.append(attr.text.strip())
        attrs_value_html = xiaomi.find_all('span', class_='attribute-value')
        attrs_value = []
        for attr in attrs_value_html:
            attrs_value.append(attr.text.strip())
        price = xiaomi.find('div', class_='price-normal selenium-price-normal').text.strip()
        link = 'https://www.euro.com.pl' + xiaomi.find('a', class_='js-save-keyword').get('href')
        # print(name, number, attrs_name, attrs_value, price, link)

        ws['A' + str(cell_number)].value = name
        ws['B' + str(cell_number)].value = number

        try:
            ws['C' + str(cell_number)].value = str(attrs_name[0])
            ws['D' + str(cell_number)].value = str(attrs_value[0])

            ws['E' + str(cell_number)].value = str(attrs_name[1])
            ws['F' + str(cell_number)].value = str(attrs_value[1])

            ws['G' + str(cell_number)].value = str(attrs_name[2])
            ws['H' + str(cell_number)].value = str(attrs_value[2])

            ws['I' + str(cell_number)].value = str(attrs_name[3])
            ws['J' + str(cell_number)].value = str(attrs_value[3])

            ws['K' + str(cell_number)].value = str(attrs_name[4])
            ws['L' + str(cell_number)].value = str(attrs_value[4])

            ws['M' + str(cell_number)].value = str(attrs_name[5])
            ws['N' + str(cell_number)].value = str(attrs_value[5])
        except IndexError:
            pass

        ws['O' + str(cell_number)].value = price
        ws['P' + str(cell_number)].value = link

        cell_number += 1
        page_counter += 1

wb.save('wynik.xlsx')
